"""
Generate a professional, formatted Excel template for MD Invoice item entry.
Run once: python create_excel_template.py
"""
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os

# ── Output path ────────────────────────────────────────────────────────────────
OUT_PATH = r"g:\MD Invoice\static\vendor\MD_Invoice_Items_Template.xlsx"

wb = openpyxl.Workbook()

# ════════════════════════════════════════════════════════════════════════════════
# Sheet 1: Items Entry
# ════════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Items Entry"

# ── Colors ─────────────────────────────────────────────────────────────────────
HEADER_BG    = "1A73E8"   # Google Blue
HEADER_FONT  = "FFFFFF"   # White
EXAMPLE_BG   = "EAF4FB"   # Light blue tint
NOTE_BG      = "FFF3CD"   # Amber note
REQUIRED_BG  = "FDE8E8"   # Light red for required
OPTIONAL_BG  = "E8F5E9"   # Light green for optional

def mkfill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def mkborder():
    thin = Side(style="thin", color="AAAAAA")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

# ── Row 1: Title banner ─────────────────────────────────────────────────────────
ws.merge_cells("A1:G1")
title_cell = ws["A1"]
title_cell.value = "MD Invoice – Product/Service Items Import Template"
title_cell.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
title_cell.fill = mkfill("0D47A1")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28

# ── Row 2: Instructions ──────────────────────────────────────────────────────────
ws.merge_cells("A2:G2")
instr = ws["A2"]
instr.value = (
    "Fill in the rows below and import using the 'Import Excel' button in the Excel Mode modal.  "
    "Red columns = required. Green columns = optional. Delete example rows before importing."
)
instr.font = Font(name="Calibri", size=10, italic=True, color="555555")
instr.fill = mkfill("EEF2FF")
instr.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[2].height = 36

# ── Row 3: Column headers ────────────────────────────────────────────────────────
COLUMNS = [
    ("A", "Description *",                    "REQUIRED – Product or service name", REQUIRED_BG),
    ("B", "HSN Code",                         "Optional – 8-digit HSN/SAC code",    OPTIONAL_BG),
    ("C", "Qty *",                            "REQUIRED – Quantity (number)",        REQUIRED_BG),
    ("D", "Rate / Unit Price *",              "REQUIRED – Price per unit in ₹",      REQUIRED_BG),
    ("E", "Discount Value",                   "Optional – 0 if no discount",         OPTIONAL_BG),
    ("F", "Discount Type",                    "Optional – 'amount' or 'percentage'", OPTIONAL_BG),
    ("G", "Tax Rate %",                       "Optional – 0 / 5 / 12 / 18 / 28",    OPTIONAL_BG),
]

WIDTHS = [50, 14, 8, 18, 16, 26, 14]

for (col_letter, label, comment, _bg), width in zip(COLUMNS, WIDTHS):
    col_idx = ord(col_letter) - ord("A") + 1
    cell = ws.cell(row=3, column=col_idx, value=label)
    cell.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    cell.fill = mkfill(HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = mkborder()
    ws.column_dimensions[col_letter].width = width
    ws.row_dimensions[3].height = 30

# ── Rows 4-6: Example data ───────────────────────────────────────────────────────
examples = [
    ["yaskawa", "85044090", 1, 0, 0, "percentage", 18],
]

for r_offset, row_data in enumerate(examples):
    row_num = 4 + r_offset
    for c_idx, value in enumerate(row_data):
        cell = ws.cell(row=row_num, column=c_idx + 1, value=value)
        cell.fill = mkfill(EXAMPLE_BG)
        cell.font = Font(name="Calibri", size=10)
        cell.border = mkborder()
        cell.alignment = Alignment(horizontal="left" if c_idx == 0 else "center", vertical="center")
    ws.row_dimensions[row_num].height = 20

# ── Rows 7-26: Empty entry rows ──────────────────────────────────────────────────
for row_num in range(7, 52):
    for c_idx in range(1, 8):
        cell = ws.cell(row=row_num, column=c_idx)
        cell.fill = mkfill("FAFAFA")
        cell.font = Font(name="Calibri", size=10)
        cell.border = mkborder()
        cell.alignment = Alignment(horizontal="left" if c_idx == 1 else "center", vertical="center")
    ws.row_dimensions[row_num].height = 20

# ── Data Validations ─────────────────────────────────────────────────────────────
# Disc. Type dropdown (col F = 6)
dv_disc = DataValidation(
    type="list", formula1='"amount,percentage"',
    allow_blank=True, showDropDown=False
)
dv_disc.sqref = "F4:F51"
dv_disc.error = "Must be 'amount' or 'percentage'"
dv_disc.errorTitle = "Invalid Discount Type"
dv_disc.prompt = "Enter: amount  OR  percentage"
dv_disc.promptTitle = "Discount Type"
ws.add_data_validation(dv_disc)

# Tax Rate dropdown (col G = 7)
dv_tax = DataValidation(
    type="list", formula1='"0,5,12,18,28"',
    allow_blank=True, showDropDown=False
)
dv_tax.sqref = "G4:G51"
dv_tax.error = "Must be 0, 5, 12, 18, or 28"
dv_tax.errorTitle = "Invalid Tax Rate"
dv_tax.prompt = "Select GST rate: 0 / 5 / 12 / 18 / 28"
dv_tax.promptTitle = "Tax Rate %"
ws.add_data_validation(dv_tax)

# Qty & Rate > 0 validation
dv_qty = DataValidation(type="decimal", operator="greaterThan", formula1="0", allow_blank=True)
dv_qty.sqref = "C4:C51"
ws.add_data_validation(dv_qty)

dv_rate = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
dv_rate.sqref = "D4:D51"
ws.add_data_validation(dv_rate)

# ── Row note at bottom ────────────────────────────────────────────────────────────
note_row = 53
ws.merge_cells(f"A{note_row}:G{note_row}")
note = ws.cell(row=note_row, column=1)
note.value = (
    "IMPORT TIPS:  Column order must match exactly.  "
    "Disc. Type must be 'amount' (flat ₹ off) or 'percentage' (% off).  "
    "Rows with empty Description are ignored.  "
    "Delete example rows (rows 4-6) before importing if not needed."
)
note.font = Font(name="Calibri", size=9, italic=True, color="664D00")
note.fill = mkfill(NOTE_BG)
note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
note.border = mkborder()
ws.row_dimensions[note_row].height = 45

# ── Freeze header rows ────────────────────────────────────────────────────────────
ws.freeze_panes = "A4"

# ════════════════════════════════════════════════════════════════════════════════
# Sheet 2: Field Reference
# ════════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Field Reference")
ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 60
ws2.column_dimensions["C"].width = 25

ws2.merge_cells("A1:C1")
ref_title = ws2["A1"]
ref_title.value = "Field Reference Guide"
ref_title.font = Font(bold=True, size=13, color="FFFFFF")
ref_title.fill = mkfill("0D47A1")
ref_title.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 24

headers2 = ["Column", "Description", "Allowed Values"]
for i, h in enumerate(headers2):
    c = ws2.cell(row=2, column=i+1, value=h)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = mkfill(HEADER_BG)
    c.border = mkborder()
    c.alignment = Alignment(horizontal="center")

ref_data = [
    ("Description *",       "Name of the product or service being sold/purchased",         "Any text (required)"),
    ("HSN Code",            "Harmonized System Nomenclature code for GST filing",           "8-digit number, e.g. 85044090"),
    ("Qty *",               "Quantity of units (required, must be > 0)",                    "Positive number, e.g. 2 or 2.5"),
    ("Rate / Unit Price *", "Price per unit in Indian Rupees (required)",                   "Positive number, e.g. 10000"),
    ("Discount Value",      "Amount or % to discount from line total",                      "Number ≥ 0, e.g. 500 or 10"),
    ("Discount Type",       "How the discount value is applied",                            "'amount' = flat ₹ | 'percentage' = %"),
    ("Tax Rate %",          "GST rate to apply on the taxable amount",                     "0 / 5 / 12 / 18 / 28"),
]

for r, (col, desc, vals) in enumerate(ref_data):
    for c_idx, val in enumerate([col, desc, vals]):
        cell = ws2.cell(row=3+r, column=c_idx+1, value=val)
        cell.font = Font(name="Calibri", size=10)
        cell.fill = mkfill("F8F9FA") if r % 2 == 0 else mkfill("FFFFFF")
        cell.border = mkborder()
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws2.row_dimensions[3+r].height = 22

# ── Save ─────────────────────────────────────────────────────────────────────────
wb.save(OUT_PATH)
size = os.path.getsize(OUT_PATH)
print(f"Template created: {OUT_PATH}")
print(f"File size: {size:,} bytes")
